import openpyxl
import pandas as pd
import json
import os
from datetime import datetime

from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller

base_dir = ""

# json타입으로 받아올 request url을 가지고 오기
def get_url(path: str, dictionary: dict) -> str:
    url = f'{base_dir}/template/page/{path}/list_json.php?'

    for key in dictionary:
        url += f'{key}={dictionary[key]}&'

    # 마지막 &는 제거
    url = url[:-1]

    return url


site = ""
user_id = "" # 지움
user_pw = "" # 지움
download_path = "" # 지움, 파일 다운로드 경로
move_path = "" # 지움, 옮길 경로
move_path = "" # 지움
target_year = "2022"
target_month = "7"

# writer = pd.ExcelWriter(f'{target_month}월 매출.xlsx', engine='openpyxl')

chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
driver_path = f'./{chrome_ver}/chromedriver.exe'
cur_driver_path = "./src/chromedriver.exe"

# 크롬 드라이버 없으면 다운로드
if not os.path.exists(cur_driver_path):
    chromedriver_autoinstaller.install(True)

    # src 폴더 생성
    if not os.path.exists('./src'):
        os.mkdir('./src')

    # 파일 옮기기 driver_path에 저장된 파일을 src폴더로 옮긴다.
    os.rename(driver_path, './src/chromedriver.exe')

    # chrome_ver 폴더 삭제
    if os.path.exists(f'./{chrome_ver}'):
        os.rmdir(f'./{chrome_ver}')

# 웹드라이버 연결
driver = webdriver.Chrome('src/chromedriver.exe')
# 사이트 접속
driver.get(site)

# 로그인
driver.find_element(By.NAME, 'user_id').send_keys(user_id)
driver.find_element(By.NAME, 'user_pw').send_keys(user_pw)
driver.find_element(By.ID, 'login_form').submit()

# 필요한 데이터를 받아올 request url 생성 후 request
request_url = get_url("report_buysell_detail",
                      {"page": '0', "start_date_input": datetime(2022, 7, 1).strftime('%Y-%m-%d'),
                       "end_date_input": datetime(2022, 7, 31).strftime('%Y-%m-%d'), "list_of_page_ch": '10000'})
driver.get(request_url)

# 받아온 response data를 json형식으로 변환
json_data = driver.find_element(By.TAG_NAME, 'body').text

# driver 종료
driver.close()

# text에서 json객체를 이용하여 dictionary로 변환
json_data = json.JSONDecoder().decode(json_data)
data = json.JSONDecoder().decode(str(json_data['return_val']))

# json 형태로 받아온 것은 영어로 값의 이름이 붙어 있기 때문에 다시 지정해 둠.
columnNames = ['매출일시', '전표번호', '고객명', '상품명', '디자이너', '단가', '수량', '할인액', '매출액', '메모(전표별)']
df = pd.read_json(json.dumps(data))
df.drop(columns=['id'], inplace=True)
df.columns = columnNames

# 합계 행 제거 -> df.loc[0]
df.drop(df.index[0], inplace=True)

# int 타입으로 변환 / ,(콤마)를 제거
df['단가'] = df['단가'].str.replace(',', '').astype(int)
df['할인액'] = df['할인액'].str.replace(',', '').astype(int)
df['매출액'] = df['매출액'].str.replace(',', '').astype(int)

df['점판'] = 0

# 점판 목록
salesList = [] # 지움


for i in df.index:
    if df['상품명'].loc[i] in salesList:
        df['점판'].loc[i] = 1


# df to excel sheet named f'{target_month} 매출'
# df.to_excel(f'{target_month}월 매출.xlsx', sheet_name=f'{target_month}월 매출', index=False, engine='openpyxl')
# storeSales = pd.DataFrame()

# f'{target_month} 매출.xlsx'에 점판 상품 목록 추가

# writer.book = openpyxl.load_workbook(f'{target_month}월 매출.xlsx')
# storeSales.to_excel(writer, sheet_name='점판')

wb = openpyxl.load_workbook('test.xlsx')
ws = wb.create_sheet(f'{target_month}월 매출')

# dataframe_to_rows
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# worksheet 복사 f'{target_month-1}월' to f'{target_month}월'
prev_manifest = wb.copy_worksheet(wb[f'{int(target_month)-1}월'])
prev_manifest.title = f'{target_month}월'

# value f'{target_month-1}월 매출' convert to f'{target_month}월 매출' in prev_manifest
prev_manifest['D4'].value = str(prev_manifest['D4'].value).replace(f'{int(target_month)-1}월 매출', f'{target_month}월 매출')
for i in range(5, 11):
    cell = 'H' + str(i)
    prev_manifest[cell].value = str(prev_manifest[cell].value).replace(f'{int(target_month) - 1}월 매출', f'{target_month}월 매출')

# remove f'{target_month}월 매출.xlsx' file
# os.remove(f'{target_month}월 매출.xlsx')
wb.save('test.xlsx')
# for i in range(len(salesList)):
#     wb.get_sheet_by_name("점판").cell(row=i+2, column=1).value = salesList[i]
# wb.save('7월 매출.xlsx')


# df to workbook sheetName=f'{target_month}월 매출'

# df.to_excel(f'{move_path}/{target_month}월 매출내역.xlsx', index=False)

# # 매출내역 페이지 이동
# for tag in driver.find_elements(By.TAG_NAME, 'div'):
#     if tag.text == '매출내역':
#         tag.click()
#         break
#
# # driver.find_element(By.ID, 'btn_month').click()
#
# # 매출기간 start_date datepicker 창 생성
# driver.find_element(By.ID, 'start_date_input').click()
#
# # 찾고자 하는 연 선택
# while driver.find_element(By.CLASS_NAME, 'ui-datepicker-year').text != target_year:
#     if int(driver.find_element(By.CLASS_NAME, 'ui-datepicker-year').text) < int(target_year):
#         driver.find_element(By.CLASS_NAME, 'ui-datepicker-next').click()
#     elif int(driver.find_element(By.CLASS_NAME, 'ui-datepicker-year').text) > int(target_year):
#         driver.find_element(By.CLASS_NAME, 'ui-datepicker-prev').click()
#
# # 찾고자 하는 달 선택
# while driver.find_element(By.CLASS_NAME, 'ui-datepicker-month').text.split('월')[0] != target_month:
#     print(driver.find_element(By.CLASS_NAME, 'ui-datepicker-month').text.split('월')[0])
#     print(driver.find_element(By.CLASS_NAME, 'ui-datepicker-month').text.split('월')[1])
#     print(target_month)
#     if int(driver.find_element(By.CLASS_NAME, 'ui-datepicker-month').text.split('월')[0]) < int(target_month):
#         driver.find_element(By.CLASS_NAME, 'ui-datepicker-next').click()
#     elif int(driver.find_element(By.CLASS_NAME, 'ui-datepicker-month').text.split('월')[0]) > int(target_month):
#         driver.find_element(By.CLASS_NAME, 'ui-datepicker-prev').click()
#
#
# # 가져올 매출의 시작 일자 선택
# for tag in driver.find_elements(By.TAG_NAME, 'a'):
#     if tag.text == '1':
#         tag.click()
#         break
#
# # 매출기간 end_date datepicker 창 생성
# driver.find_element(By.ID, 'end_date_input').click()
#
# # target_year, target_month 의 마지막 날짜 찾기
# last_date = datetime(int(target_year), int(target_month), 1) + relativedelta(months=1) - timedelta(days=1)
# print(last_date)
#
# for tag in driver.find_elements(By.TAG_NAME, 'a'):
#     if tag.text == str(last_date.day):
#         tag.click()
#         break
#
# driver.execute_script('$(\'[name="list_of_page_ch"]\').val("500")')
# driver.find_element(By.ID, 'btn_submit').click()
#
# timerTime = 1.5
#
# time.sleep(timerTime)
#
# driver.find_element(By.CLASS_NAME, 'body-menu-sub').click()
#
# print(int(driver.find_element(By.ID, 'list_count').text.split('(')[1].split(')')[0]))
# if int(driver.find_element(By.ID, 'list_count').text.split('(')[1].split(')')[0]) > 500:
#     time.sleep(timerTime * 2)
#
#     driver.execute_script('page_moves(1)')
#     time.sleep(timerTime)
#
#     driver.find_element(By.CLASS_NAME, 'body-menu-sub').click()
#
# time.sleep(timerTime)
#
# for tag in driver.find_elements(By.CLASS_NAME, 'excelBtn'):
#     tag.click()
#
# if '<i class="fal fa-table"></i><span class="body-menu-sub">엑셀다운</span>' in tag.text:
#     tag.click()
#     break


# pd_list = []


# # directory 탐색 in download_path
# for file in os.listdir(download_path):
#     if file.endswith('.xlsx'):
#
#         # 엑셀 파일 열기
#         wb = openpyxl.load_workbook(filename=download_path + '\\' + file)
#         ws = wb.active
#         print(ws.max_row)
#         print(ws.max_column)
#
#         pd_list.append(pd.read_excel(download_path + '\\' + file, engine='openpyxl'))
#
# for excel in pd_list:
#     print(excel)
#         os.rename(download_path + '\\' + file, download_path + '\\' + '매출내역.xlsx')


# with open('file.csv', 'w', encoding='UTF-8') as csvFile:
#     writer = csv.DictWriter(csvFile, fieldnames=list(data[0].keys()))
#     for row in data:
#         for key in row.keys():
#             if 'price' in key:
#                 row[key] = row[key].replace(',', '')
#             if key != 'id':
#                 csvFile.write(row[key] + ',')
#         csvFile.write('\n')

"""
홍단헤어에 있는 월관리장부.xlsx를 읽기
6월 매출 시트를 7월 매출 시트로 복사
6월 시트를 7월 시트로 복사
7월 매출 시트에 header를 제외한 모든 값들을 삭제 후 새로운 값들을 추가
7월에 있는 모든 '6월 매출'을 '7월 매출'로 변경
text.xlsx로 저장
"""