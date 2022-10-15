from datetime import datetime, timedelta
import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta

dt = datetime(2022, 7, 1)
print(dt.strftime('%Y-%m-%d'))

last_date = datetime(int("2022"), int("7"), 1) + relativedelta(months=1) - timedelta(days=1)
print(last_date.date().day)

# df = pd.read_excel('test.xlsx', sheet_name='Sheet1')
# print(df)

functionList = []
functionString = "IF(DCOLUMN=점판!$A$ROW,1,0)"
startRow = 2
점판maxRow = 33
dataMaxRow = 496
for i in range(startRow, dataMaxRow):
    functionValue = "="
    for j in range(startRow, 점판maxRow+1):
        functionValue += functionString.replace("ROW", str(j)).replace("COLUMN", str(i))
        if j != 점판maxRow:
            functionValue += "+"

    functionList.append(functionValue)
# test.txt에 저장
with open("test.txt", "w") as f:
    for i in range(len(functionList)):
        f.write(functionList[i] + "\n")

# wb = openpyxl.Workbook()
# wb.create_sheet("점판")
# wb.remove(wb.get_sheet_by_name("Sheet"))
# wb.save('test1.xlsx')
# storeSales = pd.DataFrame()
# storeSales['점판'] = functionList
# storeSales.to_excel('test1.xlsx')
salesList = [] # 지움
wb = openpyxl.load_workbook('7월 매출.xlsx')
wb.create_sheet("점판")
for i in range(len(salesList)):
    wb.get_sheet_by_name("점판").cell(row=i+2, column=1).value = salesList[i]
wb.save('7월 매출.xlsx')
